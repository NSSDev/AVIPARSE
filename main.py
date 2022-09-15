import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog, QAbstractItemView
import pandas as pd
from selenium import webdriver
import time
from tkinter import messagebox
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import os
import pyautogui

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(428, 321)
        MainWindow.setStyleSheet("background-color: rgb(53, 53, 53);")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.listWidget = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget.setGeometry(QtCore.QRect(20, 60, 391, 151))
        font = QtGui.QFont()
        font.setFamily("Rockwell Condensed")
        self.listWidget.setFont(font)
        self.listWidget.setStyleSheet("color: rgb(255, 255, 255);")
        self.listWidget.setObjectName("listWidget")
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.addItem(item)
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 40, 81, 16))
        font = QtGui.QFont()
        font.setFamily("Segoe Print")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet("color: rgb(255, 255, 255);")
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(160, 0, 111, 41))
        font = QtGui.QFont()
        font.setFamily("Niagara Solid")
        font.setPointSize(36)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(380, 260, 47, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_3.setObjectName("label_3")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(90, 230, 241, 41))
        font = QtGui.QFont()
        font.setFamily("Segoe Print")
        font.setPointSize(11)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color: rgb(0, 170, 0);\n"
"color: rgb(255, 255, 255);")
        self.pushButton.setObjectName("pushButton")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 428, 21))
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.action = QtWidgets.QAction(MainWindow)
        font = QtGui.QFont()
        font.setFamily("Segoe Print")
        self.action.setFont(font)
        self.action.setObjectName("action")
        self.action_2 = QtWidgets.QAction(MainWindow)
        font = QtGui.QFont()
        font.setFamily("Segoe Print")
        self.action_2.setFont(font)
        self.action_2.setObjectName("action_2")
        self.menu.addAction(self.action)
        self.menu.addAction(self.action_2)
        self.menubar.addAction(self.menu.menuAction())
        self.action_2.triggered.connect(self.exit_action)
        self.action.triggered.connect(self.author)
        self.pushButton.clicked.connect(self.main)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "AVIPARSE"))
        __sortingEnabled = self.listWidget.isSortingEnabled()
        self.listWidget.setSortingEnabled(False)
        item = self.listWidget.item(0)
        item.setText(_translate("MainWindow", "Транспорт и перевозки"))
        item = self.listWidget.item(1)
        item.setText(_translate("MainWindow", "Ремонт, строительство"))
        item = self.listWidget.item(2)
        item.setText(_translate("MainWindow", "Красота, здоровье"))
        item = self.listWidget.item(3)
        item.setText(_translate("MainWindow", "Ремонт техники"))
        item = self.listWidget.item(4)
        item.setText(_translate("MainWindow", "Уборка, дезинфекция"))
        item = self.listWidget.item(5)
        item.setText(_translate("MainWindow", "Оборудование, производство"))
        item = self.listWidget.item(6)
        item.setText(_translate("MainWindow", "Обучение, курсы"))
        item = self.listWidget.item(7)
        item.setText(_translate("MainWindow", "Деловые услуги"))
        item = self.listWidget.item(8)
        item.setText(_translate("MainWindow", "Установка техники"))
        item = self.listWidget.item(9)
        item.setText(_translate("MainWindow", "Мастер на час"))
        item = self.listWidget.item(10)
        item.setText(_translate("MainWindow", "Сад, участок"))
        item = self.listWidget.item(11)
        item.setText(_translate("MainWindow", "Бытовые услуги"))
        item = self.listWidget.item(12)
        item.setText(_translate("MainWindow", "Праздники, мероприятия"))
        item = self.listWidget.item(13)
        item.setText(_translate("MainWindow", "IT, интернет, телеком"))
        item = self.listWidget.item(14)
        item.setText(_translate("MainWindow", "Реклама, полиграфия"))
        item = self.listWidget.item(15)
        item.setText(_translate("MainWindow", "Главная страница"))
        self.listWidget.setSortingEnabled(__sortingEnabled)
        self.label.setText(_translate("MainWindow", "УСЛУГИ:"))
        self.label_2.setText(_translate("MainWindow", "AVIPARSE"))
        self.label_3.setText(_translate("MainWindow", "V0.0.1"))
        self.pushButton.setText(_translate("MainWindow", "Получить данные"))
        self.menu.setTitle(_translate("MainWindow", "Меню"))
        self.action.setText(_translate("MainWindow", "Автор"))
        self.action_2.setText(_translate("MainWindow", "Закрыть"))


    def author(self):
        messagebox.showinfo(title="AVIPARSE v0.0.1", message="Разработчик программы:\nTelegram: @HasimNas")

    def exit_action(self):
        MainWindow.close()

    def select_file_directory(self):
        self.file_path = QFileDialog.getExistingDirectory(None, caption='Выберите папку для сохранения файла',
                                                          directory=os.getcwd())
    def choose_category(self):
        self.select = self.listWidget.currentRow()
        if self.select == 0:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/transport_perevozki-ASgBAgICAUSYC8SfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 1:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/remont_stroitelstvo-ASgBAgICAUSYC8CfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 2:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/krasota_zdorove-ASgBAgICAUSYC6qfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 3:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/remont_i_obsluzhivanie_tehniki-ASgBAgICAUSYC7T3AQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 4:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/uborka_klining-ASgBAgICAUSYC7L3AQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 5:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/oborudovanie_proizvodstvo-ASgBAgICAUSYC7SfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 6:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/obuchenie_kursy-ASgBAgICAUSYC7afAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 7:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/delovye_uslugi-ASgBAgICAUSYC7KfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self .select == 8:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/ustanovka_tehniki-ASgBAgICAUSYC9r1AQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 9:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/master_na_chas-ASgBAgICAUSYC7zvAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 10:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/sad_blagoustroystvo-ASgBAgICAUSYC8KfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 11:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/bytovye_uslugi-ASgBAgICAUSYC7CfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 12:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/prazdniki_meropriyatiya-ASgBAgICAUSYC7yfAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 13:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/it_internet_telekom-ASgBAgICAUSYC6afAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 14:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug/reklama_poligrafiya-ASgBAgICAUSYC76fAQ?context=H4sIAAAAAAAA_0q0MrKqLraysFJKK8rPDUhMT1WyLrYyNLNSKipNKspMTizJLwrPTElPLQGLG1gplSQWpaeWwFQaWCkpWdcCAgAA__-OZLTGRgAAAA&p=2"
        if self.select == 15:
            self.url = "https://www.avito.ru/rossiya/predlozheniya_uslug?cd=1&isVertical=1"


    def collect_info(self):
        for r, d, f in os.walk("C:\\"):
                for files in f:
                    if files == "geckodriver.exe":
                        self.driver_path = os.path.join(r, files)
        options = webdriver.FirefoxOptions()
        options.add_argument("--headless")
        self.driver = webdriver.Firefox(executable_path=self.driver_path, options=options)
        self.driver.maximize_window()
        counter = 2

        while counter != 100:
            try:
                self.driver.get(url=f"{self.url[:-1]}{counter}")
                counter += 1
                time.sleep(7)
                if self.select == 15:
                    items = self.driver.find_elements(By.CLASS_NAME, "styles-list-M1e9B")
                else:
                    items = self.driver.find_elements(By.CLASS_NAME, "items-items-kAJAg")
                for item in items:
                    links = item.find_element(By.TAG_NAME, "a").get_attribute('href')
                    self.driver.get(links)
                    time.sleep(2)
                    id = links[-10:]
                    self.driver.find_element(By.XPATH,
                                        "//button[@class='button-button-eBrUW button-button_phone-_Yo3v button-button_card-AkthM button-button-CmK9a button-size-s-r9SeD button-success-_ytiQ']").click()
                    time.sleep(2)
                    image = self.driver.find_element(By.XPATH, "//img[@alt='phone']").get_attribute('src')
                    self.driver.get(image)
                    if counter == 30:
                        time.sleep(30)
                    os.chdir(self.file_path)
                    if os.path.exists(f"{self.file_path}\\contacts.xlsx"):
                        fn = "contacts.xlsx"
                        wb = load_workbook(fn)
                        ws = wb['ContactsInfo']
                        ws.append(["ID"])
                        for i in ws['A']:
                            if id in i.value:
                                break
                        else:
                            self.driver.save_screenshot(f'{self.file_path}/{id}.png')
                            ws.append([id])
                            wb.save(fn)
                            print(f"[INFO] Добавлено {id} {counter}")
                    else:
                        pd = openpyxl.Workbook()
                        pd_sheet = pd["Sheet"]
                        pd_sheet.title = "ContactsInfo"
                        ps = pd.active
                        ps.append(["ID"])
                        pd.save(filename=f"{self.file_path}/contacts.xlsx")

            except Exception as e:
                print(e)

    def main(self):
        self.choose_category() #проверка на выбор подкатегории
        self.select_file_directory() #выбор пути для сохранения результатов
        if not self.file_path: #если путь не выбран
            messagebox.showerror(title="AVIPARSE",message="Вы не выбрали папку для сохранения файла !")
            return
        else:
            self.collect_info() #основной метод для парсинга данных
            messagebox.showinfo(title="AVIPARSE",message="Программа закончила свою работу") #по завершению закрываем selenium
            self.driver.close()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
